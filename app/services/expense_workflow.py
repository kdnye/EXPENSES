"""Expense workflow helpers for spreadsheet lookups, uploads, and dispatch."""

from __future__ import annotations

import csv
import io
import os
import uuid
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import List, Sequence, Tuple

import openpyxl
import paramiko
from flask import current_app
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from app.models import ExpenseReport


class ExpenseReferenceDataError(RuntimeError):
    """Raised when the runtime expense reference workbook cannot be consumed.

    Inputs:
        message: Operator-facing guidance that explains why workbook access
            failed and how to remediate the runtime deployment.

    Outputs:
        A domain-specific exception consumed by route handlers to avoid
        unhandled 500 errors for workbook-related failures.

    External dependencies:
        None.
    """


@dataclass(frozen=True)
class GLAccountOption:
    """Searchable GL account option sourced from the reference workbook."""

    account: str
    label: str


@dataclass(frozen=True)
class ExpenseLineDecision:
    """Decision metadata collected for a single expense line review."""

    line_id: int
    status: str
    comment: str


@lru_cache(maxsize=1)
def _workbook_path() -> Path:
    """Return the canonical workbook path used to mirror spreadsheet workflows."""

    return Path(current_app.root_path).parent / "expense_report_template.xlsx"


def _load_reference_workbook(*, required_sheet: str) -> openpyxl.Workbook:
    """Load the shared workbook and return it when the required sheet exists.

    Inputs:
        required_sheet: Worksheet name that must be available for the caller,
            such as ``GL Accounts`` or ``Data List``.

    Outputs:
        An open :class:`openpyxl.workbook.workbook.Workbook` object in
        read-only mode. Callers must close it after reading values.

    External dependencies:
        * Calls :func:`app.services.expense_workflow._workbook_path` to resolve
          the expected runtime path.
        * Calls :func:`openpyxl.load_workbook` to read spreadsheet data.
    """

    workbook_path = _workbook_path()
    expected_sheets = ("GL Accounts", "Data List")

    try:
        workbook = openpyxl.load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
        )
    except (
        FileNotFoundError,
        openpyxl.utils.exceptions.InvalidFileException,
        OSError,
    ) as exc:
        raise ExpenseReferenceDataError(
            "Expense reference workbook could not be loaded. "
            f"Expected file: '{workbook_path}'. "
            f"Required sheets: {', '.join(expected_sheets)}. "
            "Ensure the workbook exists and is a valid .xlsx file on the "
            "application host."
        ) from exc

    try:
        workbook[required_sheet]
    except KeyError as exc:
        workbook.close()
        raise ExpenseReferenceDataError(
            "Expense reference workbook is missing required sheet data. "
            f"Expected file: '{workbook_path}'. "
            f"Required sheets: {', '.join(expected_sheets)}. "
            "Verify the deployed workbook matches the template structure."
        ) from exc

    return workbook


@lru_cache(maxsize=1)
def load_gl_accounts() -> Tuple[GLAccountOption, ...]:
    """Read GL account values from the workbook ``GL Accounts`` sheet."""

    workbook = _load_reference_workbook(required_sheet="GL Accounts")
    sheet = workbook["GL Accounts"]
    values: List[GLAccountOption] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        account = str(row[0] or "").strip()
        label = str(row[1] or "").strip()
        if not account:
            continue
        display = f"{account} - {label}" if label else account
        values.append(GLAccountOption(account=account, label=display))
    workbook.close()
    return tuple(values)


@lru_cache(maxsize=1)
def load_expense_types() -> Tuple[str, ...]:
    """Read standardized expense types from the workbook ``Data List`` sheet."""

    workbook = _load_reference_workbook(required_sheet="Data List")
    sheet = workbook["Data List"]
    values: List[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        candidate = str(row[0] or "").strip()
        if candidate:
            values.append(candidate)
    workbook.close()
    return tuple(values)


def upload_receipt_to_cloud_storage(
    file_storage: FileStorage,
    *,
    report_id: int,
    line_index: int,
) -> str:
    """Upload a receipt image to GCS and return its URL.

    This function calls ``google.cloud.storage.Client`` when the
    ``EXPENSE_RECEIPT_BUCKET`` configuration is present.
    """

    if not file_storage or not file_storage.filename:
        return ""

    bucket_name = (current_app.config.get("EXPENSE_RECEIPT_BUCKET") or "").strip()
    if not bucket_name:
        return ""

    from google.cloud import storage  # Imported lazily to keep startup fast.

    safe_name = secure_filename(file_storage.filename)
    extension = Path(safe_name).suffix.lower()
    unique_name = (
        f"expense-receipts/{report_id}/{line_index}-{uuid.uuid4().hex}{extension}"
    )

    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(unique_name)
    blob.upload_from_file(
        file_storage.stream,
        content_type=file_storage.content_type or "application/octet-stream",
    )
    blob.make_public()
    return blob.public_url


def parse_line_review_form_data(
    report: ExpenseReport,
    form_data: dict[str, str],
) -> Tuple[ExpenseLineDecision, ...]:
    """Collect line review decisions from a submitted form payload.

    Inputs:
        report: The :class:`app.models.ExpenseReport` with lines to review.
        form_data: Raw form data keyed by ``line_status_<id>`` and
            ``line_comment_<id>``.

    Outputs:
        A tuple of :class:`ExpenseLineDecision` values aligned to each report
        line.

    External dependencies:
        Reads :attr:`app.models.ExpenseReport.lines` to build the result.
    """

    decisions: List[ExpenseLineDecision] = []
    for line in report.lines:
        status_key = f"line_status_{line.id}"
        comment_key = f"line_comment_{line.id}"
        decisions.append(
            ExpenseLineDecision(
                line_id=line.id,
                status=(form_data.get(status_key) or "").strip(),
                comment=(form_data.get(comment_key) or "").strip(),
            )
        )
    return tuple(decisions)


def apply_line_review_decisions(
    report: ExpenseReport,
    *,
    decisions: Sequence[ExpenseLineDecision],
) -> Tuple[str, str]:
    """Apply line-level review decisions to a report and its expenses.

    Inputs:
        report: The :class:`app.models.ExpenseReport` instance being reviewed.
        decisions: Sequence of line review decisions for each report line.

    Outputs:
        A two-item tuple containing the flash message and category that the
        caller should present to the user interface.

    External dependencies:
        * Updates :class:`app.models.ExpenseReport` and ``ExpenseLine`` objects
          in memory. Callers must commit changes via ``app.models.db.session``.
    """

    decision_map = {decision.line_id: decision for decision in decisions}
    missing_lines = [line.id for line in report.lines if line.id not in decision_map]
    if missing_lines:
        raise ValueError("Select approve or reject for each expense line.")

    any_rejected = False
    for line in report.lines:
        decision = decision_map.get(line.id)
        if decision is None:
            raise ValueError("Select approve or reject for each expense line.")

        normalized_status = decision.status.strip().lower()
        if normalized_status not in {"approve", "reject"}:
            raise ValueError("Select approve or reject for each expense line.")

        if normalized_status == "reject":
            if not decision.comment:
                raise ValueError(
                    "Add a rejection comment for each rejected expense line."
                )
            line.status = "Rejected"
            line.rejection_comment = decision.comment
            any_rejected = True
        else:
            line.status = "Approved"
            line.rejection_comment = None

    if any_rejected:
        report.status = "Draft"
        report.rejection_comment = "One or more expense lines were rejected. Review line comments and resubmit."
        return (
            "Report updated with rejected lines and returned to draft status.",
            "info",
        )

    report.status = "Pending Upload"
    report.rejection_comment = None
    return "Report approved and queued for NetSuite upload.", "success"


def format_pending_reports_csv(reports: Sequence[ExpenseReport]) -> str:
    """Serialize ``Pending Upload`` reports into a single NetSuite-ready CSV."""

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "report_id",
            "employee_email",
            "supervisor_email",
            "expense_date",
            "expense_type",
            "gl_account",
            "vendor",
            "description",
            "amount",
            "receipt_url",
        ]
    )

    for report in reports:
        for line in report.lines:
            writer.writerow(
                [
                    report.id,
                    getattr(report.employee, "email", ""),
                    getattr(report.supervisor, "email", ""),
                    line.date.isoformat(),
                    line.expense_type,
                    line.gl_account,
                    line.vendor,
                    line.description or "",
                    f"{line.amount:.2f}",
                    line.receipt_url or "",
                ]
            )

    return output.getvalue()


def dispatch_csv_via_sftp(payload: str, *, filename: str) -> None:
    """Transmit a generated expense export to the configured NetSuite SFTP host."""

    host = (current_app.config.get("NETSUITE_SFTP_HOST") or "").strip()
    username = (current_app.config.get("NETSUITE_SFTP_USERNAME") or "").strip()
    password = (current_app.config.get("NETSUITE_SFTP_PASSWORD") or "").strip()
    remote_dir = (current_app.config.get("NETSUITE_SFTP_DIRECTORY") or "/").strip()
    port = int(current_app.config.get("NETSUITE_SFTP_PORT", 22))

    if not host or not username or not password:
        raise ValueError("NetSuite SFTP credentials are not fully configured.")

    transport = paramiko.Transport((host, port))
    try:
        transport.connect(username=username, password=password)
        client = paramiko.SFTPClient.from_transport(transport)
        remote_path = f"{remote_dir.rstrip('/')}/{filename}"
        with client.file(remote_path, "w") as remote_file:
            remote_file.write(payload)
    finally:
        transport.close()
